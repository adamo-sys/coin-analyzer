"""Tests for shared per-session workbook and WANT_LIST context."""

import os
import tempfile
import unittest

from openpyxl import Workbook

from acquisition_workflow import AcquisitionWorkflow
from buy_advisor import BuyAdvisor
from coin_collection import CoinCollection, CoinItem
from focused_collection_intelligence import CandidateItem, FocusedCollectionIntelligenceEngine, MatchStatus
from session_context import SessionContext


CORE_HEADERS = [
    "Item",
    "Type",
    "Year",
    "Denomination",
    "Variety",
    "Grade",
    "Certifier",
    "Certification #",
    "Purchase Price",
    "Estimated Value",
    "Running Total",
    "Status",
    "Liquidity Score",
    "Notes",
    "Acquired From",
    "Date Acquired",
    "Source",
    "Numista #",
    "Bullion Value CAD",
    "Dealer Bid CAD",
    "Retail Value CAD",
    "Priority",
    "Silver?",
    "ASW oz",
    "Portfolio Category",
]

WANT_HEADERS = [
    "Target Coin",
    "Priority",
    "Target Grade",
    "Budget",
    "Why Wanted",
    "Status",
]


def make_item(item_id, country, denomination, year, grade):
    return CoinItem(
        id=item_id,
        image_path="",
        country=country,
        denomination=denomination,
        year=year,
        grade=grade,
        notes="",
        date_added="2026-06-16",
    )


def create_workbook(path, want_rows=None, include_inventory=True):
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    if include_inventory:
        core = wb.create_sheet("CORE_RAW")
        core.append(CORE_HEADERS)
        core.append([
            "Newfoundland - 50 Cents",
            "COIN",
            1901,
            "50 cents",
            "KM# 6",
            "VF-20",
            "RAW",
            "",
            20,
            85,
            "",
            "KEEP",
            7,
            "Session context test row",
            "Dealer",
            "2026-06-16",
            "Manual",
            "",
            "",
            "",
            "",
            "High",
            "Yes",
            "0.35",
            "Newfoundland",
        ])

    want = wb.create_sheet("WANT_LIST")
    want.append(WANT_HEADERS)
    for row in want_rows or []:
        want.append(row)

    wb.save(path)


class TestSessionContext(unittest.TestCase):
    """Verify shared context loading, reuse, and clearing."""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = os.path.join(self.temp_dir.name, "legacy_context.xlsx")
        self.collection_path = os.path.join(self.temp_dir.name, "collection.json")
        self.collection = CoinCollection(self.collection_path)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_empty_session_context(self):
        context = SessionContext()

        self.assertEqual(context.loaded_collection_workbook_path, "")
        self.assertEqual(context.loaded_collection_item_count, 0)
        self.assertEqual(context.want_list_count, 0)
        self.assertEqual(context.get_want_list_intents(), [])
        self.assertIn("none loaded", context.format_status_line())

    def test_load_collection_context_successfully(self):
        create_workbook(self.workbook_path)
        context = SessionContext()

        result = context.load_collection_context(self.workbook_path, [])

        self.assertTrue(result.success)
        self.assertEqual(context.loaded_collection_workbook_path, self.workbook_path)
        self.assertEqual(context.loaded_collection.rows_found, 1)
        self.assertEqual(context.loaded_collection_item_count, 1)
        self.assertEqual(context.load_status, "Collection context loaded")

    def test_load_want_list_context_successfully(self):
        create_workbook(self.workbook_path, want_rows=[[
            "Newfoundland 50 cents 1904",
            "High",
            "VF-20",
            150,
            "Session target",
            "Active",
        ]])
        context = SessionContext()

        result = context.load_want_list_context(self.workbook_path, [])

        self.assertTrue(result.success)
        self.assertEqual(context.want_list_count, 1)
        self.assertEqual(len(context.get_want_list_intents()), 1)
        self.assertEqual(context.get_want_list_intents()[0].target_coin, "Newfoundland 50 cents 1904")

    def test_missing_workbook_handled_gracefully(self):
        context = SessionContext()

        result = context.load_workbook_context(os.path.join(self.temp_dir.name, "missing.xlsx"), [])

        self.assertFalse(result.success)
        self.assertIn("Workbook not found", result.errors[0])
        self.assertEqual(context.get_want_list_intents(), [])

    def test_invalid_workbook_handled_gracefully(self):
        invalid_path = os.path.join(self.temp_dir.name, "invalid.xlsx")
        with open(invalid_path, "w", encoding="utf-8") as handle:
            handle.write("not an excel workbook")
        context = SessionContext()

        result = context.load_workbook_context(invalid_path, [])

        self.assertFalse(result.success)
        self.assertTrue(result.errors)
        self.assertEqual(context.get_want_list_intents(), [])

    def test_do_i_own_this_uses_shared_want_list_context(self):
        create_workbook(self.workbook_path, want_rows=[[
            "Newfoundland 50 cents 1904",
            "High",
            "VF-20",
            150,
            "Session target",
            "Active",
        ]])
        context = SessionContext()
        context.load_want_list_context(self.workbook_path, [])

        engine = FocusedCollectionIntelligenceEngine([], context.get_want_list_intents())
        result = engine.analyze_candidate(CandidateItem(
            country="Newfoundland",
            denomination="50 cents",
            year="1904",
            grade="VF-20",
        ))

        self.assertEqual(result.match_status, MatchStatus.WANT_LIST_MATCH)
        self.assertEqual(result.want_list_status, "ON_WANT_LIST")

    def test_acquisition_workflow_uses_shared_want_list_context(self):
        create_workbook(self.workbook_path, want_rows=[[
            "Newfoundland 50 cents 1904",
            "High",
            "VF-20",
            150,
            "Session target",
            "Active",
        ]])
        context = SessionContext()
        context.load_want_list_context(self.workbook_path, [])

        decision = AcquisitionWorkflow([], context.get_want_list_intents()).evaluate(CandidateItem(
            country="Newfoundland",
            denomination="50 cents",
            year="1904",
            grade="VF-20",
            asking_price=125.0,
            certifier="PCGS",
        ))

        self.assertEqual(decision.collection_intelligence_status, "WANT_LIST_MATCH")
        self.assertEqual(decision.recommendation, "BUY")

    def test_buy_advisor_can_access_shared_context(self):
        create_workbook(self.workbook_path, want_rows=[[
            "Newfoundland 50 cents 1904",
            "High",
            "VF-20",
            150,
            "Session target",
            "Active",
        ]])
        context = SessionContext()
        context.load_want_list_context(self.workbook_path, [])
        advisor = BuyAdvisor(self.collection, staged_want_list_intents=context.get_want_list_intents())

        recommendation = advisor.advise(
            "Newfoundland",
            "50 cents",
            "1904",
            grade="VF-20",
            asking_price=125.0,
            estimated_market_value=150.0,
        )

        self.assertIn("+50 Explicit WANT_LIST target", recommendation.collection_intelligence_factors)
        self.assertGreater(recommendation.collection_impact_score, 0)

    def test_clear_session_context_resets_state(self):
        create_workbook(self.workbook_path, want_rows=[[
            "Canada 1 cent 1920",
            "Medium",
            "VF-20",
            25,
            "Session target",
            "Active",
        ]])
        context = SessionContext()
        context.load_workbook_context(self.workbook_path, [])

        result = context.clear()

        self.assertTrue(result.success)
        self.assertEqual(context.loaded_collection_workbook_path, "")
        self.assertEqual(context.want_list_count, 0)
        self.assertEqual(context.get_want_list_intents(), [])

    def test_existing_manual_workflow_still_works(self):
        engine = FocusedCollectionIntelligenceEngine([
            make_item("1", "Canada", "1 cent", "1967", "VF-30")
        ])

        result = engine.analyze_candidate(CandidateItem(
            country="Canada",
            denomination="1c",
            year="1967",
            grade="VF-30",
        ))

        self.assertEqual(result.match_status, MatchStatus.SAME_GRADE_DUPLICATE)
        self.assertEqual(result.recommendation, "PASS")


if __name__ == "__main__":
    unittest.main()
