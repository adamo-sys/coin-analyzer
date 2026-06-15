"""
Tests for the safe legacy portfolio workbook preview importer.
"""

import os
import csv
import shutil
import tempfile
import unittest

from openpyxl import Workbook

from coin_collection import CoinCollection
from legacy_portfolio_importer import (
    LegacyPortfolioImporter,
    LegacyWantListIntent,
    export_import_summary_csv,
)


FIXTURE_COLLECTION = os.path.join(
    os.path.dirname(__file__),
    "test_data",
    "sample_collection.json",
)


HEADERS_CORE_RAW = [
    "Item",
    "Type",
    "Year",
    "Denomination",
    "Variety",
    "Grade",
    "Certifier",
    "Certification #",
    "Purchase Price",
    "Estimated Value",
    "Running Total",
    "Status",
    "Liquidity Score",
    "Notes",
    "Acquired From",
    "Date Acquired",
    "Source",
    "Numista #",
    "Bullion Value CAD",
    "Dealer Bid CAD",
    "Retail Value CAD",
    "Priority",
    "Silver?",
    "ASW oz",
    "Portfolio Category",
    "Disposition",
    "Eye Appeal",
    "Liquidity",
    "Attribution Confidence",
    "Rarity",
    "Acquisition Source",
    "Submission Candidate",
    "Expected Grade",
    "Upside Potential",
    "Collection Tier",
]

HEADERS_SLABS = HEADERS_CORE_RAW[:25]
HEADERS_WANT_LIST = [
    "Target Coin",
    "Priority",
    "Target Grade",
    "Budget",
    "Why Wanted",
    "Status",
]


class TestLegacyPortfolioImporter(unittest.TestCase):
    """Verify Phase 1 legacy workbook preview behavior."""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.collection_path = os.path.join(self.temp_dir.name, "collection.json")
        self.workbook_path = os.path.join(self.temp_dir.name, "legacy_portfolio.xlsx")
        shutil.copy(FIXTURE_COLLECTION, self.collection_path)
        self.collection_before = self._read_collection_file()
        self.collection = CoinCollection(self.collection_path)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_preview_stages_core_raw_and_slabs_without_saving_collection(self):
        self._create_workbook(
            core_rows=[
                [
                    "Canada - 1 Cent",
                    "COIN",
                    1966,
                    "1 cent",
                    "KM# 59",
                    "VF-20",
                    "RAW",
                    "",
                    0.25,
                    1.0,
                    "",
                    "KEEP",
                    3,
                    "Duplicate by Numista",
                    "Existing fixture",
                    "2025-01-01",
                    "Numista",
                    "N# 1001",
                    "",
                    "",
                    "",
                    "High",
                    "No",
                    "",
                    "Canadian base",
                ],
                [
                    "Newfoundland - 50 Cents",
                    "COIN",
                    1900,
                    "50 cents",
                    "KM# 6",
                    "F-12",
                    "RAW",
                    "",
                    20,
                    85,
                    "",
                    "KEEP",
                    7,
                    "Priority Newfoundland row",
                    "Dealer",
                    "2025-02-01",
                    "Manual",
                    "",
                    "",
                    "",
                    "",
                    "High",
                    "Yes",
                    "0.35",
                    "Newfoundland",
                ],
                [
                    "Canada - Test Note",
                    "NOTE",
                    1937,
                    "1 dollar",
                    "",
                    "VF",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "KEEP",
                    "",
                    "Unsupported type",
                    "",
                    "",
                    "Manual",
                    "",
                ],
            ],
            slab_rows=[
                [
                    "1901 Canada 5 cents ICCS EF40",
                    "COIN",
                    1901,
                    "5 cents",
                    "KM# 13",
                    "EF-40",
                    "ICCS",
                    "XDI 093",
                    30,
                    40,
                    "",
                    "KEEP",
                    6,
                    "Slabbed coin",
                    "Existing",
                    "2025-03-01",
                    "Manual",
                    "",
                    3.47,
                    28,
                    40,
                    "Medium",
                    "Yes",
                    "0.0346",
                    "Canadian silver",
                ]
            ],
        )

        importer = LegacyPortfolioImporter(self.collection.items)
        summary = importer.preview_workbook(self.workbook_path)

        self.assertEqual(summary.rows_found, 4)
        self.assertEqual(summary.items_importable, 2)
        self.assertEqual(summary.duplicates_detected, 1)
        self.assertEqual(summary.rows_skipped, 1)
        self.assertIn("Rows found: 4", summary.format_summary())
        self.assertEqual(self._read_collection_file(), self.collection_before)

        staged = {item.coin_item.title: item for item in summary.staged_items}
        self.assertIn("Newfoundland - 50 Cents", staged)
        self.assertIn("1901 Canada 5 cents ICCS EF40", staged)
        self.assertEqual(staged["Newfoundland - 50 Cents"].coin_item.country, "Newfoundland")
        self.assertEqual(staged["Newfoundland - 50 Cents"].coin_item.estimate_cad, 85.0)
        self.assertEqual(
            staged["1901 Canada 5 cents ICCS EF40"].metadata["Certification #"],
            "XDI 093",
        )

        duplicate = summary.duplicate_items[0]
        self.assertEqual(duplicate.duplicate_of, "fixture_canada_cent_1966")
        self.assertIn("Numista", duplicate.duplicate_reason)
        self.assertEqual(summary.skipped_rows[0].reason, "Unsupported Type value: NOTE")

    def test_missing_inventory_sheet_is_reported_as_warning(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "CORE_RAW"
        worksheet.append(HEADERS_CORE_RAW)
        workbook.save(self.workbook_path)

        importer = LegacyPortfolioImporter(self.collection.items)
        summary = importer.preview_workbook(self.workbook_path)

        self.assertEqual(summary.rows_found, 0)
        self.assertEqual(summary.items_importable, 0)
        self.assertIn("Missing required sheet: SLABS", summary.warnings)
        self.assertEqual(self._read_collection_file(), self.collection_before)

    def test_export_import_summary_csv_writes_preview_report(self):
        self._create_workbook(
            core_rows=[
                [
                    "Canada - 1 Cent",
                    "COIN",
                    1966,
                    "1 cent",
                    "KM# 59",
                    "VF-20",
                    "RAW",
                    "",
                    0.25,
                    1.0,
                    "",
                    "KEEP",
                    3,
                    "Duplicate by Numista",
                    "Existing fixture",
                    "2025-01-01",
                    "Numista",
                    "N# 1001",
                ],
                [
                    "1900",
                    "COIN",
                    1900,
                    "token",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "KEEP",
                    "",
                    "Ambiguous country",
                    "",
                    "",
                    "Manual",
                    "",
                ],
            ],
            slab_rows=[],
        )
        summary = LegacyPortfolioImporter(self.collection.items).preview_workbook(
            self.workbook_path
        )
        output_path = os.path.join(self.temp_dir.name, "portfolio_preview.csv")

        self.assertTrue(export_import_summary_csv(summary, output_path))

        with open(output_path, "r", encoding="utf-8") as handle:
            rows = list(csv.reader(handle))

        self.assertIn(["Summary", "Rows Found", "2"], rows)
        self.assertIn(["Summary", "Importable Items", "1"], rows)
        self.assertIn(["Summary", "Duplicates", "1"], rows)
        self.assertTrue(any(len(row) > 3 and row[0] == "Staged" and row[3] == "1900" for row in rows))
        self.assertTrue(any(len(row) > 10 and row[0] == "Duplicate" and row[10] == "fixture_canada_cent_1966" for row in rows))
        self.assertTrue(any(row and row[0] == "Warning" for row in rows))
        self.assertEqual(self._read_collection_file(), self.collection_before)

    def test_empty_want_list_sheet_returns_empty_preview(self):
        self._create_want_list_workbook([])

        preview = LegacyPortfolioImporter(self.collection.items).preview_want_list(
            self.workbook_path
        )

        self.assertEqual(preview.rows_found, 0)
        self.assertEqual(preview.intents_staged, 0)
        self.assertEqual(preview.rows_skipped, 0)
        self.assertEqual(preview.staged_intents, [])
        self.assertEqual(self._read_collection_file(), self.collection_before)

    def test_populated_want_list_sheet_stages_acquisition_intent(self):
        self._create_want_list_workbook(
            [
                [
                    "Newfoundland 50 Cents 1904",
                    "High",
                    "VF-20",
                    "125.50",
                    "Priority date-run target",
                    "Active",
                ],
                [
                    "Canada 1859 Large Cent Narrow 9",
                    "Medium",
                    "EF-40",
                    200,
                    "Variety attribution target",
                    "Watching",
                ],
            ]
        )

        preview = LegacyPortfolioImporter(self.collection.items).preview_want_list(
            self.workbook_path
        )

        self.assertEqual(preview.rows_found, 2)
        self.assertEqual(preview.intents_staged, 2)
        self.assertEqual(preview.rows_skipped, 0)
        self.assertIsInstance(preview.staged_intents[0], LegacyWantListIntent)
        self.assertEqual(preview.staged_intents[0].target_coin, "Newfoundland 50 Cents 1904")
        self.assertEqual(preview.staged_intents[0].priority, "High")
        self.assertEqual(preview.staged_intents[0].target_grade, "VF-20")
        self.assertEqual(preview.staged_intents[0].budget, 125.50)
        self.assertEqual(preview.staged_intents[0].why_wanted, "Priority date-run target")
        self.assertEqual(preview.staged_intents[0].status, "Active")
        self.assertNotEqual(preview.staged_intents[0].legacy_id, "")
        self.assertEqual(self._read_collection_file(), self.collection_before)

    def test_invalid_want_list_rows_are_skipped_with_warnings(self):
        self._create_want_list_workbook(
            [
                ["", "High", "VF-20", 100, "Missing target", "Active"],
                [
                    "Canada 10 Cents 1911",
                    "High",
                    "VF-20",
                    "not money",
                    "Budget parse warning",
                    "Active",
                ],
            ]
        )

        preview = LegacyPortfolioImporter(self.collection.items).preview_want_list(
            self.workbook_path
        )

        self.assertEqual(preview.rows_found, 2)
        self.assertEqual(preview.intents_staged, 1)
        self.assertEqual(preview.rows_skipped, 1)
        self.assertEqual(preview.skipped_rows[0].reason, "Missing Target Coin")
        self.assertIn("Budget could not be parsed", preview.warnings[0])
        self.assertEqual(self._read_collection_file(), self.collection_before)

    def test_want_list_priority_order_is_deterministic(self):
        self._create_want_list_workbook(
            [
                ["Canada 25 Cents 1912", "Low", "VF-20", 50, "", "Active"],
                ["Newfoundland 20 Cents 1888", "High", "VF-20", 150, "", "Active"],
                ["Canada 10 Cents 1911", "High", "VF-20", 100, "", "Active"],
                ["Canada 1859 Large Cent", "2", "VF-20", 200, "", "Active"],
            ]
        )

        preview = LegacyPortfolioImporter(self.collection.items).preview_want_list(
            self.workbook_path
        )

        self.assertEqual(
            [intent.target_coin for intent in preview.staged_intents],
            [
                "Canada 10 Cents 1911",
                "Newfoundland 20 Cents 1888",
                "Canada 25 Cents 1912",
                "Canada 1859 Large Cent",
            ],
        )
        self.assertEqual(
            [intent.priority_score for intent in preview.staged_intents],
            [75, 75, 25, 2],
        )

    def _create_workbook(self, core_rows, slab_rows):
        workbook = Workbook()
        core_sheet = workbook.active
        core_sheet.title = "CORE_RAW"
        core_sheet.append(HEADERS_CORE_RAW)
        for row in core_rows:
            core_sheet.append(row)

        slab_sheet = workbook.create_sheet("SLABS")
        slab_sheet.append(HEADERS_SLABS)
        for row in slab_rows:
            slab_sheet.append(row)

        workbook.save(self.workbook_path)

    def _create_want_list_workbook(self, want_list_rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "WANT_LIST"
        worksheet.append(HEADERS_WANT_LIST)
        for row in want_list_rows:
            worksheet.append(row)
        workbook.save(self.workbook_path)

    def _read_collection_file(self):
        with open(self.collection_path, "r", encoding="utf-8") as handle:
            return handle.read()


if __name__ == "__main__":
    unittest.main()
