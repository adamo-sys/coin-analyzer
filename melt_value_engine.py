"""Melt Value Engine for calculating silver coin melt values."""

import json
import os
import time
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from collection_intelligence import SILVER_DENOMINATION_TERMS


@dataclass
class ASWReferenceEntry:
    """ASW (Actual Silver Weight) reference entry from workbook."""
    country_series: str
    denomination: str
    year_range: str
    asw_oz: float
    notes: str


@dataclass
class MeltValueResult:
    """Melt value calculation result."""
    coin_country: str
    coin_denomination: str
    coin_year: str
    asw_oz: float
    spot_price_cad: float
    melt_value_cad: float
    is_silver: bool
    confidence: str  # HIGH, MEDIUM, LOW, NONE
    source: str  # WORKBOOK, CALCULATED, ESTIMATED
    workbook_bullion_value: Optional[float] = None  # Preserved reference value
    spot_price_warning: Optional[str] = None  # Warning if spot price is stale or failed


class SpotPriceProvider(ABC):
    """Abstract base class for silver spot price providers."""
    
    @abstractmethod
    def get_spot_price(self) -> float:
        """Get current silver spot price in CAD per ounce."""
        pass
    
    @abstractmethod
    def refresh_spot_price(self) -> tuple[float, Optional[str]]:
        """
        Refresh spot price from source.
        
        Returns:
            Tuple of (spot_price, warning_message)
        """
        pass


class ManualSpotPriceProvider(SpotPriceProvider):
    """Manual spot price provider with optional override."""
    
    def __init__(self, default_spot_price_cad: float = 35.0):
        """
        Initialize with default spot price.
        
        Args:
            default_spot_price_cad: Default silver spot price in CAD per ounce
        """
        self.default_spot_price_cad = default_spot_price_cad
        self.manual_spot_price_cad: Optional[float] = None
    
    def set_manual_spot_price(self, spot_price_cad: float) -> None:
        """Set manual spot price override."""
        self.manual_spot_price_cad = spot_price_cad
    
    def get_spot_price(self) -> float:
        """Get current spot price (manual override or default)."""
        return self.manual_spot_price_cad if self.manual_spot_price_cad is not None else self.default_spot_price_cad
    
    def refresh_spot_price(self) -> tuple[float, Optional[str]]:
        """Refresh spot price (no-op for manual provider)."""
        return self.get_spot_price(), None
    
    def reset_to_default(self) -> None:
        """Reset to default spot price."""
        self.manual_spot_price_cad = None


class ApiSpotPriceProvider(SpotPriceProvider):
    """API-based spot price provider with caching and fallback."""
    
    CACHE_FILE = "data/silver_spot_cache.json"
    CACHE_DURATION_HOURS = 24
    
    def __init__(self, default_spot_price_cad: float = 35.0):
        """
        Initialize with default spot price.
        
        Args:
            default_spot_price_cad: Default silver spot price in CAD per ounce
        """
        self.default_spot_price_cad = default_spot_price_cad
        self.manual_spot_price_cad: Optional[float] = None
        self._cached_price: Optional[float] = None
        self._cache_timestamp: Optional[float] = None
        self._load_cache()
    
    def set_manual_spot_price(self, spot_price_cad: float) -> None:
        """Set manual spot price override."""
        self.manual_spot_price_cad = spot_price_cad
    
    def get_spot_price(self) -> float:
        """Get current spot price (manual override, cached, or default)."""
        if self.manual_spot_price_cad is not None:
            return self.manual_spot_price_cad
        
        if self._cached_price is not None and self._is_cache_valid():
            return self._cached_price
        
        return self.default_spot_price_cad
    
    def refresh_spot_price(self) -> tuple[float, Optional[str]]:
        """
        Refresh spot price from API.
        
        Returns:
            Tuple of (spot_price, warning_message)
        """
        try:
            # Try to fetch from API
            new_price = self._fetch_from_api()
            if new_price:
                self._cached_price = new_price
                self._cache_timestamp = time.time()
                self._save_cache()
                return new_price, None
            else:
                # API returned no data, use cached or default
                return self._get_fallback_price(), "API returned no data, using cached/default price"
        except Exception as e:
            # API failed, use cached or default
            return self._get_fallback_price(), f"API failed: {str(e)}, using cached/default price"
    
    def _fetch_from_api(self) -> Optional[float]:
        """
        Fetch spot price from API.
        
        Returns:
            Spot price in CAD per ounce or None if fetch fails
        """
        # Placeholder for actual API implementation
        # For now, return None to simulate API failure
        # In production, this would call a real API like:
        # - metals-api.com
        # - silverprice.org
        # - kitco.com
        return None
    
    def _get_fallback_price(self) -> float:
        """Get fallback price (cached or default)."""
        if self._cached_price is not None:
            return self._cached_price
        return self.default_spot_price_cad
    
    def _is_cache_valid(self) -> bool:
        """Check if cache is still valid (within 24 hours)."""
        if self._cache_timestamp is None:
            return False
        
        cache_age = time.time() - self._cache_timestamp
        return cache_age < (self.CACHE_DURATION_HOURS * 3600)
    
    def _load_cache(self) -> None:
        """Load cached spot price from file."""
        try:
            if os.path.exists(self.CACHE_FILE):
                with open(self.CACHE_FILE, 'r') as f:
                    cache_data = json.load(f)
                    self._cached_price = cache_data.get('price')
                    self._cache_timestamp = cache_data.get('timestamp')
        except Exception:
            # If cache loading fails, start fresh
            self._cached_price = None
            self._cache_timestamp = None
    
    def _save_cache(self) -> None:
        """Save cached spot price to file."""
        try:
            os.makedirs(os.path.dirname(self.CACHE_FILE), exist_ok=True)
            with open(self.CACHE_FILE, 'w') as f:
                json.dump({
                    'price': self._cached_price,
                    'timestamp': self._cache_timestamp
                }, f)
        except Exception:
            # If cache saving fails, continue without cache
            pass
    
    def reset_to_default(self) -> None:
        """Reset to default spot price."""
        self.manual_spot_price_cad = None
        self._cached_price = None
        self._cache_timestamp = None


class ASWReferenceLoader:
    """Loads and caches ASW reference data from workbook."""
    
    def __init__(self):
        """Initialize ASW reference loader."""
        self._asw_cache: Dict[str, ASWReferenceEntry] = {}
        self._loaded = False
    
    def load_from_workbook(self, workbook_path: str) -> List[ASWReferenceEntry]:
        """
        Load ASW reference data from workbook.
        
        Args:
            workbook_path: Path to legacy workbook
            
        Returns:
            List of ASW reference entries
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(workbook_path, read_only=True)
            
            if "ASW_REFERENCE" not in wb.sheetnames:
                return []
            
            sheet = wb["ASW_REFERENCE"]
            entries = []
            
            # Skip header row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:  # Skip empty rows
                    continue
                
                entry = ASWReferenceEntry(
                    country_series=str(row[0]) if row[0] else "",
                    denomination=str(row[1]) if row[1] else "",
                    year_range=str(row[2]) if row[2] else "",
                    asw_oz=float(row[3]) if row[3] else 0.0,
                    notes=str(row[4]) if row[4] else ""
                )
                entries.append(entry)
                self._cache_entry(entry)
            
            self._loaded = True
            wb.close()
            return entries
            
        except Exception as e:
            # If workbook loading fails, return empty list
            return []
    
    def _cache_entry(self, entry: ASWReferenceEntry) -> None:
        """Cache entry for quick lookup."""
        key = self._make_key(entry.country_series, entry.denomination)
        self._asw_cache[key] = entry
        self._loaded = True
    
    def _make_key(self, country_series: str, denomination: str) -> str:
        """Create lookup key."""
        return f"{country_series.lower()}|{denomination.lower()}"
    
    def find_asw_entry(self, country: str, denomination: str) -> Optional[ASWReferenceEntry]:
        """
        Find ASW entry for country/denomination.
        
        Args:
            country: Coin country
            denomination: Coin denomination
            
        Returns:
            ASW reference entry or None
        """
        key = self._make_key(country, denomination)
        return self._asw_cache.get(key)
    
    def is_loaded(self) -> bool:
        """Check if ASW data has been loaded."""
        return self._loaded
    
    def clear_cache(self) -> None:
        """Clear cached ASW data."""
        self._asw_cache.clear()
        self._loaded = False


class MeltValueEngine:
    """Calculates melt values for silver coins using ASW reference data."""
    
    def __init__(self, asw_loader: ASWReferenceLoader, spot_provider: SpotPriceProvider):
        """
        Initialize melt value engine.
        
        Args:
            asw_loader: ASW reference loader
            spot_provider: Silver spot price provider
        """
        self.asw_loader = asw_loader
        self.spot_provider = spot_provider
    
    def calculate_melt_value(
        self,
        country: str,
        denomination: str,
        year: str,
        manual_asw_oz: Optional[float] = None,
        workbook_bullion_value: Optional[float] = None,
        refresh_spot: bool = False
    ) -> MeltValueResult:
        """
        Calculate melt value for a coin.
        
        Args:
            country: Coin country
            denomination: Coin denomination
            year: Coin year
            manual_asw_oz: Manual ASW override (optional)
            workbook_bullion_value: Workbook bullion value for reference (optional)
            refresh_spot: Whether to refresh spot price from API
            
        Returns:
            MeltValueResult with calculation details
        """
        # Refresh spot price if requested
        spot_price_warning = None
        if refresh_spot:
            spot_price, spot_price_warning = self.spot_provider.refresh_spot_price()
        else:
            spot_price = self.spot_provider.get_spot_price()
        
        # Determine if coin is silver
        is_silver = self._is_silver_coin(country, denomination)
        
        # Get ASW value
        if manual_asw_oz is not None:
            asw_oz = manual_asw_oz
            source = "CALCULATED"
            confidence = "HIGH"
        else:
            asw_entry = self.asw_loader.find_asw_entry(country, denomination)
            if asw_entry:
                asw_oz = asw_entry.asw_oz
                source = "WORKBOOK"
                confidence = "HIGH"
            else:
                # Estimate ASW for silver coins
                asw_oz = self._estimate_asw(denomination)
                source = "ESTIMATED"
                confidence = "LOW" if asw_oz > 0 else "NONE"
        
        # Calculate melt value
        melt_value = asw_oz * spot_price if is_silver else 0.0
        
        return MeltValueResult(
            coin_country=country,
            coin_denomination=denomination,
            coin_year=year,
            asw_oz=asw_oz,
            spot_price_cad=spot_price,
            melt_value_cad=melt_value,
            is_silver=is_silver,
            confidence=confidence,
            source=source,
            workbook_bullion_value=workbook_bullion_value,
            spot_price_warning=spot_price_warning
        )
    
    def _is_silver_coin(self, country: str, denomination: str) -> bool:
        """
        Determine if coin is silver based on denomination.
        
        Args:
            country: Coin country
            denomination: Coin denomination
            
        Returns:
            True if silver coin, False otherwise
        """
        country_lower = country.lower()
        denom_lower = denomination.lower()
        
        # Canadian silver denominations
        if "canada" in country_lower:
            return any(term in denom_lower for term in SILVER_DENOMINATION_TERMS)
        
        # Newfoundland silver (5c, 10c, 20c, 50c were silver)
        if "newfoundland" in country_lower:
            silver_denoms = ["5 cent", "10 cent", "20 cent", "50 cent", "5c", "10c", "20c", "50c"]
            return any(d in denom_lower for d in silver_denoms)
        
        return False
    
    def _estimate_asw(self, denomination: str) -> float:
        """
        Estimate ASW for silver coin when reference data unavailable.
        
        Args:
            denomination: Coin denomination
            
        Returns:
            Estimated ASW in ounces
        """
        denom_lower = denomination.lower()
        
        # Rough ASW estimates for common Canadian silver denominations
        asw_estimates = {
            "5 cent": 0.0588,
            "10 cent": 0.0588,
            "25 cent": 0.0588,
            "50 cent": 0.0588,
            "dollar": 0.0588,
            "5c": 0.0588,
            "10c": 0.0588,
            "25c": 0.0588,
            "50c": 0.0588,
            "$1": 0.0588,
            "quarter": 0.0588,
            "dime": 0.0588,
        }
        
        for key, asw in asw_estimates.items():
            if key in denom_lower:
                return asw
        
        return 0.0
