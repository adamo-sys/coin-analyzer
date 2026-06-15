"""
Safe preview importer for Adam's legacy portfolio workbook.

Phase 1 intentionally stages CORE_RAW and SLABS rows only. It does not write to
data/collection.json or provide a merge path; permanent imports must be added as
a separate confirmed workflow later.
"""

from __future__ import annotations

import re
import csv
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from coin_collection import CoinItem


INVENTORY_SHEETS = ("CORE_RAW", "SLABS")
WANT_LIST_SHEET = "WANT_LIST"

WANT_LIST_HEADERS = (
    "Target Coin",
    "Priority",
    "Target Grade",
    "Budget",
    "Why Wanted",
    "Status",
)

REQUIRED_HEADERS = (
    "Item",
    "Type",
    "Year",
    "Denomination",
    "Variety",
    "Grade",
    "Certifier",
    "Certification #",
    "Purchase Price",
    "Estimated Value",
    "Status",
    "Notes",
    "Date Acquired",
    "Source",
    "Numista #",
)

FUTURE_METADATA_HEADERS = (
    "Certifier",
    "Certification #",
    "Purchase Price",
    "Running Total",
    "Status",
    "Liquidity Score",
    "Acquired From",
    "Date Acquired",
    "Source",
    "Bullion Value CAD",
    "Dealer Bid CAD",
    "Retail Value CAD",
    "Priority",
    "Silver?",
    "ASW oz",
    "Portfolio Category",
    "Disposition",
    "Eye Appeal",
    "Liquidity",
    "Attribution Confidence",
    "Rarity",
    "Acquisition Source",
    "Submission Candidate",
    "Expected Grade",
    "Upside Potential",
    "Collection Tier",
)

KNOWN_COUNTRIES = (
    "Newfoundland",
    "Canada",
    "Argentina",
    "Australia",
    "United States",
    "Great Britain",
    "United Kingdom",
    "England",
    "France",
    "Germany",
    "Mexico",
    "India",
    "Japan",
    "China",
    "Italy",
    "Spain",
    "Portugal",
    "Switzerland",
    "Netherlands",
    "Belgium",
)


@dataclass
class LegacyPortfolioStagedItem:
    """A workbook row mapped to a CoinItem plus review metadata."""

    sheet_name: str
    row_number: int
    legacy_id: str
    coin_item: CoinItem
    metadata: Dict[str, Any] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    duplicate_of: Optional[str] = None
    duplicate_reason: str = ""

    @property
    def is_duplicate(self) -> bool:
        return bool(self.duplicate_of)


@dataclass
class LegacyPortfolioSkippedRow:
    """A workbook row that could not safely become a staged item."""

    sheet_name: str
    row_number: int
    reason: str


@dataclass
class LegacyWantListIntent:
    """A workbook WANT_LIST row staged as acquisition intent, not inventory."""

    sheet_name: str
    row_number: int
    legacy_id: str
    target_coin: str
    priority: str = ""
    target_grade: str = ""
    budget: float = 0.0
    why_wanted: str = ""
    status: str = ""
    priority_score: int = 0
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
            "legacy_id": self.legacy_id,
            "target_coin": self.target_coin,
            "priority": self.priority,
            "target_grade": self.target_grade,
            "budget": self.budget,
            "why_wanted": self.why_wanted,
            "status": self.status,
            "priority_score": self.priority_score,
            "warnings": list(self.warnings),
        }


@dataclass
class LegacyPortfolioImportSummary:
    """Reviewable result of a legacy workbook preview."""

    rows_found: int = 0
    items_importable: int = 0
    duplicates_detected: int = 0
    rows_skipped: int = 0
    warnings: List[str] = field(default_factory=list)
    staged_items: List[LegacyPortfolioStagedItem] = field(default_factory=list)
    duplicate_items: List[LegacyPortfolioStagedItem] = field(default_factory=list)
    skipped_rows: List[LegacyPortfolioSkippedRow] = field(default_factory=list)

    def add_warning(self, warning: str) -> None:
        self.warnings.append(warning)

    def format_summary(self) -> str:
        """Return a compact text summary suitable for an import preview."""

        lines = [
            "Legacy Portfolio Import Preview",
            f"Rows found: {self.rows_found}",
            f"Items importable: {self.items_importable}",
            f"Duplicates detected: {self.duplicates_detected}",
            f"Rows skipped: {self.rows_skipped}",
            f"Warnings: {len(self.warnings)}",
        ]
        if self.warnings:
            lines.append("")
            lines.append("Warnings:")
            lines.extend(f"- {warning}" for warning in self.warnings)
        return "\n".join(lines)

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable summary for tests or future GUI preview code."""

        return {
            "rows_found": self.rows_found,
            "items_importable": self.items_importable,
            "duplicates_detected": self.duplicates_detected,
            "rows_skipped": self.rows_skipped,
            "warnings": list(self.warnings),
            "staged_items": [item.legacy_id for item in self.staged_items],
            "duplicate_items": [item.legacy_id for item in self.duplicate_items],
            "skipped_rows": [
                {
                    "sheet_name": row.sheet_name,
                    "row_number": row.row_number,
                    "reason": row.reason,
                }
                for row in self.skipped_rows
            ],
        }


@dataclass
class LegacyWantListPreview:
    """Reviewable result of a workbook WANT_LIST preview."""

    rows_found: int = 0
    intents_staged: int = 0
    rows_skipped: int = 0
    warnings: List[str] = field(default_factory=list)
    staged_intents: List[LegacyWantListIntent] = field(default_factory=list)
    skipped_rows: List[LegacyPortfolioSkippedRow] = field(default_factory=list)

    def add_warning(self, warning: str) -> None:
        self.warnings.append(warning)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "rows_found": self.rows_found,
            "intents_staged": self.intents_staged,
            "rows_skipped": self.rows_skipped,
            "warnings": list(self.warnings),
            "staged_intents": [intent.to_dict() for intent in self.staged_intents],
            "skipped_rows": [
                {
                    "sheet_name": row.sheet_name,
                    "row_number": row.row_number,
                    "reason": row.reason,
                }
                for row in self.skipped_rows
            ],
        }

    def sorted_intents(self) -> List[LegacyWantListIntent]:
        """Return staged intents in deterministic priority order."""

        return sorted(
            self.staged_intents,
            key=lambda intent: (
                -intent.priority_score,
                intent.target_coin.lower(),
                intent.row_number,
            ),
        )


class LegacyPortfolioImporter:
    """Build a safe preview of CORE_RAW and SLABS workbook rows."""

    def __init__(self, existing_items: Optional[Sequence[CoinItem]] = None):
        self.existing_items = list(existing_items or [])
        self._numista_index, self._identity_index = self._build_duplicate_indexes(
            self.existing_items
        )

    def preview_workbook(self, workbook_path: str) -> LegacyPortfolioImportSummary:
        """
        Parse CORE_RAW and SLABS into staged records without modifying app data.

        The returned summary separates importable records, likely duplicates, and
        skipped rows. No collection save or merge operation is performed here.
        """

        summary = LegacyPortfolioImportSummary()
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)

        for sheet_name in INVENTORY_SHEETS:
            if sheet_name not in workbook.sheetnames:
                summary.add_warning(f"Missing required sheet: {sheet_name}")
                continue

            sheet = workbook[sheet_name]
            header_map = self._header_map(sheet)
            missing_headers = [
                header for header in REQUIRED_HEADERS if header not in header_map
            ]
            if missing_headers:
                summary.add_warning(
                    f"{sheet_name} missing required headers: "
                    + ", ".join(missing_headers)
                )
                continue

            for row_number, row_values in self._iter_data_rows(sheet, header_map):
                summary.rows_found += 1
                staged, skip_reason = self._stage_row(
                    sheet_name, row_number, row_values
                )
                if skip_reason:
                    summary.rows_skipped += 1
                    summary.skipped_rows.append(
                        LegacyPortfolioSkippedRow(sheet_name, row_number, skip_reason)
                    )
                    continue

                duplicate_id, duplicate_reason = self._find_duplicate(
                    staged.coin_item, staged.metadata
                )
                if duplicate_id:
                    staged.duplicate_of = duplicate_id
                    staged.duplicate_reason = duplicate_reason
                    summary.duplicates_detected += 1
                    summary.duplicate_items.append(staged)
                else:
                    summary.items_importable += 1
                    summary.staged_items.append(staged)

                summary.warnings.extend(staged.warnings)

        workbook.close()
        return summary

    def preview_want_list(self, workbook_path: str) -> LegacyWantListPreview:
        """
        Parse WANT_LIST rows into acquisition-intent staging data.

        This is read-only and deliberately separate from inventory import. It does
        not create CoinItem objects and does not save collection data.
        """

        preview = LegacyWantListPreview()
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            if WANT_LIST_SHEET not in workbook.sheetnames:
                preview.add_warning(f"Missing optional sheet: {WANT_LIST_SHEET}")
                return preview

            sheet = workbook[WANT_LIST_SHEET]
            header_map = self._header_map(sheet)
            missing_headers = [
                header for header in WANT_LIST_HEADERS if header not in header_map
            ]
            if missing_headers:
                preview.add_warning(
                    f"{WANT_LIST_SHEET} missing required headers: "
                    + ", ".join(missing_headers)
                )
                return preview

            for row_number, row_values in self._iter_data_rows(sheet, header_map):
                preview.rows_found += 1
                intent, skip_reason = self._stage_want_list_row(row_number, row_values)
                if skip_reason:
                    preview.rows_skipped += 1
                    preview.skipped_rows.append(
                        LegacyPortfolioSkippedRow(
                            WANT_LIST_SHEET, row_number, skip_reason
                        )
                    )
                    continue

                preview.intents_staged += 1
                preview.staged_intents.append(intent)
                preview.warnings.extend(intent.warnings)

            preview.staged_intents = preview.sorted_intents()
            return preview
        finally:
            workbook.close()

    def _stage_want_list_row(
        self, row_number: int, row: Dict[str, Any]
    ) -> Tuple[Optional[LegacyWantListIntent], str]:
        target_coin = _clean_text(row.get("Target Coin"))
        if not target_coin:
            return None, "Missing Target Coin"

        priority = _clean_text(row.get("Priority"))
        target_grade = _clean_text(row.get("Target Grade"))
        budget = _to_float(row.get("Budget"))
        why_wanted = _clean_text(row.get("Why Wanted"))
        status = _clean_text(row.get("Status"))
        warnings: List[str] = []
        if row.get("Budget") not in ("", None) and budget <= 0:
            warnings.append(
                f"{WANT_LIST_SHEET} row {row_number}: Budget could not be parsed"
            )

        return (
            LegacyWantListIntent(
                sheet_name=WANT_LIST_SHEET,
                row_number=row_number,
                legacy_id=_want_list_legacy_id(row_number, target_coin),
                target_coin=target_coin,
                priority=priority,
                target_grade=target_grade,
                budget=budget,
                why_wanted=why_wanted,
                status=status,
                priority_score=_priority_score(priority),
                warnings=warnings,
            ),
            "",
        )

    def _stage_row(
        self, sheet_name: str, row_number: int, row: Dict[str, Any]
    ) -> Tuple[Optional[LegacyPortfolioStagedItem], str]:
        item_text = _clean_text(row.get("Item"))
        row_type = _clean_text(row.get("Type")).upper()
        year = _normalize_year(row.get("Year"))
        denomination = _clean_text(row.get("Denomination"))
        variety = _clean_text(row.get("Variety"))

        if row_type and row_type != "COIN":
            return None, f"Unsupported Type value: {row_type}"
        if not item_text and not (year and denomination):
            return None, "Missing item identity"

        warnings: List[str] = []
        country = _infer_country(item_text)
        if not country:
            warnings.append(
                f"{sheet_name} row {row_number}: country could not be inferred"
            )

        numista_n = _normalize_numista(row.get("Numista #"))
        source = _clean_text(row.get("Source"))
        notes = _clean_text(row.get("Notes"))
        comments = _build_comments(sheet_name, row)
        legacy_id = _legacy_id(sheet_name, row_number, item_text, year, denomination)
        metadata = {
            header: _clean_metadata_value(row.get(header))
            for header in FUTURE_METADATA_HEADERS
            if header in row and _clean_metadata_value(row.get(header)) not in ("", None)
        }
        metadata["legacy_sheet"] = sheet_name
        metadata["legacy_row"] = row_number

        coin_item = CoinItem(
            id=legacy_id,
            image_path="",
            country=country,
            denomination=denomination,
            year=year,
            grade=_clean_text(row.get("Grade")),
            notes=notes,
            date_added=_date_to_string(row.get("Date Acquired")),
            auto_detected=False,
            detection_confidence=0.0,
            issuer=country,
            currency="",
            face_value="",
            reference=variety,
            numista_n=numista_n,
            title=item_text,
            quantity=1,
            estimate_cad=_to_float(row.get("Estimated Value")),
            comments=comments,
            from_numista=bool(numista_n or "NUMISTA" in source.upper()),
        )

        return (
            LegacyPortfolioStagedItem(
                sheet_name=sheet_name,
                row_number=row_number,
                legacy_id=legacy_id,
                coin_item=coin_item,
                metadata=metadata,
                warnings=warnings,
            ),
            "",
        )

    def _find_duplicate(
        self, coin_item: CoinItem, metadata: Dict[str, Any]
    ) -> Tuple[Optional[str], str]:
        numista_n = _normalize_numista(coin_item.numista_n)
        if numista_n and numista_n in self._numista_index:
            return self._numista_index[numista_n], f"Numista # {numista_n}"

        identity = _identity_key(
            coin_item.country,
            coin_item.denomination,
            coin_item.year,
            coin_item.reference,
        )
        if identity and identity in self._identity_index:
            return self._identity_index[identity], "country/denomination/year/variety"

        cert_number = _normalize_cert_number(metadata.get("Certification #"))
        if cert_number:
            for existing in self.existing_items:
                existing_text = f"{existing.notes} {existing.comments}"
                if cert_number and cert_number in _normalize_cert_number(existing_text):
                    return existing.id, f"certification # {cert_number}"

        return None, ""

    @staticmethod
    def _build_duplicate_indexes(
        existing_items: Iterable[CoinItem],
    ) -> Tuple[Dict[str, str], Dict[Tuple[str, str, str, str], str]]:
        numista_index: Dict[str, str] = {}
        identity_index: Dict[Tuple[str, str, str, str], str] = {}

        for item in existing_items:
            numista_n = _normalize_numista(item.numista_n)
            if numista_n:
                numista_index.setdefault(numista_n, item.id)

            identity = _identity_key(
                item.country, item.denomination, item.year, item.reference
            )
            if identity:
                identity_index.setdefault(identity, item.id)

        return numista_index, identity_index

    @staticmethod
    def _header_map(sheet: Any) -> Dict[str, int]:
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            return {
                _clean_text(value): index
                for index, value in enumerate(row)
                if _clean_text(value)
            }
        return {}

    @staticmethod
    def _iter_data_rows(sheet: Any, header_map: Dict[str, int]):
        header_names = list(header_map.keys())
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(value not in ("", None) for value in values):
                continue
            yield row_number, {
                header: values[index] if index < len(values) else None
                for header, index in header_map.items()
                if header in header_names
            }


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _clean_metadata_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def _normalize_year(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    match = re.search(r"\d{3,4}", text)
    return match.group(0) if match else text


def _normalize_numista(value: Any) -> str:
    text = _clean_text(value).upper()
    if not text:
        return ""
    text = text.replace("N#", "").replace("NUMISTA", "")
    return re.sub(r"[^0-9A-Z]", "", text)


def _normalize_cert_number(value: Any) -> str:
    return re.sub(r"[^0-9A-Z]", "", _clean_text(value).upper())


def _to_float(value: Any) -> float:
    if value in ("", None):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(text) if text else 0.0
    except ValueError:
        return 0.0


def _date_to_string(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _clean_text(value)


def _infer_country(item_text: str) -> str:
    if not item_text:
        return ""

    if " - " in item_text:
        prefix = item_text.split(" - ", 1)[0].strip()
        if not re.fullmatch(r"\d{3,4}", prefix):
            return prefix

    searchable = f" {item_text.lower()} "
    for country in KNOWN_COUNTRIES:
        if f" {country.lower()} " in searchable:
            return country

    without_year = re.sub(r"^\s*\d{3,4}\s+", "", item_text).strip()
    if without_year:
        first_token = re.split(r"\s+", without_year, maxsplit=1)[0]
        if first_token and not first_token[0].isdigit():
            return first_token
    return ""


def _legacy_id(
    sheet_name: str, row_number: int, item_text: str, year: str, denomination: str
) -> str:
    slug_source = "_".join(part for part in (item_text, year, denomination) if part)
    slug = re.sub(r"[^a-z0-9]+", "_", slug_source.lower()).strip("_")
    if len(slug) > 60:
        slug = slug[:60].rstrip("_")
    return f"legacy_{sheet_name.lower()}_{row_number}_{slug or 'row'}"


def _want_list_legacy_id(row_number: int, target_coin: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", target_coin.lower()).strip("_")
    if len(slug) > 60:
        slug = slug[:60].rstrip("_")
    return f"legacy_want_list_{row_number}_{slug or 'target'}"


def _priority_score(priority: Any) -> int:
    text = _clean_text(priority).lower()
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        pass

    if text in ("urgent", "critical"):
        return 100
    if text in ("high", "h"):
        return 75
    if text in ("medium", "med", "m"):
        return 50
    if text in ("low", "l"):
        return 25
    return 10


def _build_comments(sheet_name: str, row: Dict[str, Any]) -> str:
    parts = [f"Legacy source: {sheet_name}"]
    for header in ("Certifier", "Certification #", "Source"):
        value = _clean_text(row.get(header))
        if value:
            parts.append(f"{header}: {value}")
    return "; ".join(parts)


def _identity_key(
    country: Any, denomination: Any, year: Any, reference: Any
) -> Tuple[str, str, str, str]:
    country_text = _normalize_identity_part(country)
    denomination_text = _normalize_identity_part(denomination)
    year_text = _normalize_year(year)
    reference_text = _normalize_identity_part(reference)
    if not (country_text and denomination_text and year_text):
        return ()
    return (country_text, denomination_text, year_text, reference_text)


def _normalize_identity_part(value: Any) -> str:
    text = _clean_text(value).lower()
    text = text.replace("cents", "cent").replace("centavos", "centavo")
    return re.sub(r"\s+", " ", text).strip()


def export_import_summary_csv(
    summary: LegacyPortfolioImportSummary, output_path: str
) -> bool:
    """Export a preview summary, staged rows, duplicates, skipped rows, and warnings."""

    try:
        with open(output_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Section", "Field", "Value"])
            writer.writerow(["Summary", "Rows Found", summary.rows_found])
            writer.writerow(["Summary", "Importable Items", summary.items_importable])
            writer.writerow(["Summary", "Duplicates", summary.duplicates_detected])
            writer.writerow(["Summary", "Skipped Rows", summary.rows_skipped])
            writer.writerow(["Summary", "Warnings", len(summary.warnings)])
            writer.writerow([])

            writer.writerow(
                [
                    "Section",
                    "Sheet",
                    "Row",
                    "Title",
                    "Country",
                    "Denomination",
                    "Year",
                    "Grade",
                    "Estimate CAD",
                    "Numista #",
                    "Duplicate Of",
                    "Reason",
                ]
            )

            for staged in summary.staged_items:
                writer.writerow(_staged_item_report_row("Staged", staged))

            for duplicate in summary.duplicate_items:
                writer.writerow(_staged_item_report_row("Duplicate", duplicate))

            for skipped in summary.skipped_rows:
                writer.writerow(
                    [
                        "Skipped",
                        skipped.sheet_name,
                        skipped.row_number,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        skipped.reason,
                    ]
                )

            for warning in summary.warnings:
                writer.writerow(["Warning", "", "", "", "", "", "", "", "", "", "", warning])

        return True
    except Exception:
        return False


def _staged_item_report_row(section: str, staged: LegacyPortfolioStagedItem) -> List[Any]:
    item = staged.coin_item
    return [
        section,
        staged.sheet_name,
        staged.row_number,
        item.title,
        item.country,
        item.denomination,
        item.year,
        item.grade,
        item.estimate_cad,
        item.numista_n,
        staged.duplicate_of or "",
        staged.duplicate_reason or "; ".join(staged.warnings),
    ]


def export_want_list_preview_csv(
    preview: LegacyWantListPreview, output_path: str
) -> bool:
    """Export staged WANT_LIST acquisition-intent preview data to CSV."""

    try:
        with open(output_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Section", "Field", "Value"])
            writer.writerow(["Summary", "Total Intents Found", preview.rows_found])
            writer.writerow(["Summary", "Valid Intents", preview.intents_staged])
            writer.writerow(["Summary", "Skipped Rows", preview.rows_skipped])
            writer.writerow(["Summary", "Warnings", len(preview.warnings)])
            writer.writerow([])

            writer.writerow(
                [
                    "Section",
                    "Sheet",
                    "Row",
                    "Target Coin",
                    "Priority",
                    "Priority Score",
                    "Target Grade",
                    "Budget",
                    "Why Wanted",
                    "Status",
                    "Reason",
                ]
            )
            for intent in preview.staged_intents:
                writer.writerow(
                    [
                        "Intent",
                        intent.sheet_name,
                        intent.row_number,
                        intent.target_coin,
                        intent.priority,
                        intent.priority_score,
                        intent.target_grade,
                        intent.budget,
                        intent.why_wanted,
                        intent.status,
                        "; ".join(intent.warnings),
                    ]
                )

            for skipped in preview.skipped_rows:
                writer.writerow(
                    [
                        "Skipped",
                        skipped.sheet_name,
                        skipped.row_number,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        skipped.reason,
                    ]
                )

            for warning in preview.warnings:
                writer.writerow(["Warning", "", "", "", "", "", "", "", "", "", warning])

        return True
    except Exception:
        return False
